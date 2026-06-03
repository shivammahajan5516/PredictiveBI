# loader.py — Loads real Excel survey data (default or uploaded)
# Reads the actual raw response rows from the survey file.

import pandas as pd
import numpy as np
from collections import Counter
import statistics
import io

DEFAULT_PATH = "data/survey_responses.xlsx"

# ── Column indices (0-based) in the Excel file ────────────────────────────────
COL = dict(
    timestamp=0, role=1, exp=2, method=3, org=4, sector=5,
    q6=6,q7=7,q8=8,q9=9,q10=10,
    q11=11,q12=12,q13=13,q14=14,q15=15,
    q16=16,q17=17,q18=18,q19=19,
    q20=20,q21=21,
    q22=22,q23=23,q24=24,q25=25,
    q26=26, q27=27, q28=28,
)

BI_COLS      = [6,7,8,9,10]
PA_COLS      = [11,12,13,14,15]
TRUST_COLS   = [16,17,18,19]
ETHICS_COLS  = [20,21]
SUCCESS_COLS = [22,23,24,25]

Q_LABELS = {
    "Q6":  "Organisation uses BI dashboards to monitor performance",
    "Q7":  "BI dashboards provide timely and accurate insights",
    "Q8":  "BI tools make it easier to identify risks",
    "Q9":  "BI dashboards present data in a clear format",
    "Q10": "BI tools enhanced data-driven decision-making",
    "Q11": "Organisation uses PA tools for planning",
    "Q12": "PA tools use historical data for forecasting",
    "Q13": "Regularly use PA indicators to anticipate delays",
    "Q14": "PA tools help identify scope creep / constraints",
    "Q15": "PA tools are regularly updated",
    "Q16": "I trust PA tool outputs for project decisions",
    "Q17": "PA outputs are explained clearly enough",
    "Q18": "BI/PA tools are easy to use",
    "Q19": "Managers are sufficiently trained to use PA outputs",
    "Q20": "Concerned about bias in PA evaluating performance",
    "Q21": "Concerned about how personal/team data is used",
    "Q22": "Projects are typically delivered on time",
    "Q23": "BI/PA tools contributed to improved on-time delivery",
    "Q24": "PA tools helped reduce post-release defects",
    "Q25": "PA + BI integration improved overall performance",
}


def _parse_excel(source):
    """Load workbook from path or BytesIO and return (headers, data_rows)."""
    import openpyxl
    if isinstance(source, (str, bytes)):
        wb = openpyxl.load_workbook(source)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(source.read()) if hasattr(source,'read') else source)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]


def _parse_csv(source):
    """Load CSV and return (headers, data_rows)."""
    df = pd.read_csv(source)
    rows = [tuple(df.columns)] + [tuple(r) for r in df.itertuples(index=False, name=None)]
    return rows[0], rows[1:]


def load_survey(source=None):
    """
    Load survey data from source (uploaded file) or default Excel.
    Returns a SurveyData object.
    """
    if source is None:
        headers, raw = _parse_excel(DEFAULT_PATH)
        label = f"Dissertation survey dataset (n = {len(raw)})"
    else:
        name = getattr(source, 'name', '')
        if name.endswith('.csv'):
            headers, raw = _parse_csv(source)
        else:
            headers, raw = _parse_excel(source)
        label = f"Uploaded file: {name} (n = {len(raw)})"

    return SurveyData(raw, label)


class SurveyData:
    """Holds all computed statistics for one dataset."""

    def __init__(self, raw, label):
        self.raw   = raw
        self.label = label
        self.n     = len(raw)
        self._compute()

    # ── Internal helpers ──────────────────────────────────────────────────────
    def _val(self, col_idx):
        """Return list of floats for one Likert column."""
        out = []
        for r in self.raw:
            try: out.append(float(r[col_idx]))
            except: pass
        return out

    def _str_col(self, col_idx):
        return [str(r[col_idx]).strip() for r in self.raw if r[col_idx] and str(r[col_idx]).strip()]

    @staticmethod
    def _mean(vals):  return round(statistics.mean(vals), 3) if vals else 0
    @staticmethod
    def _agree(vals): return round(sum(1 for v in vals if v>=4)/len(vals)*100,1) if vals else 0
    @staticmethod
    def _neutral(v):  return round(sum(1 for x in v if x==3)/len(v)*100,1) if v else 0
    @staticmethod
    def _disagree(v): return round(sum(1 for x in v if x<=2)/len(v)*100,1) if v else 0

    @staticmethod
    def _pearson(x, y):
        n = len(x)
        if n < 2: return 0
        mx,my = sum(x)/n, sum(y)/n
        num = sum((xi-mx)*(yi-my) for xi,yi in zip(x,y))
        den = (sum((xi-mx)**2 for xi in x)*sum((yi-my)**2 for yi in y))**0.5
        return round(num/den,3) if den else 0

    @staticmethod
    def _cronbach(matrix):
        k = len(matrix[0])
        item_vars = [statistics.variance([row[i] for row in matrix]) for i in range(k)]
        total = [sum(row) for row in matrix]
        total_var = statistics.variance(total) if len(total)>1 else 1
        return round((k/(k-1))*(1-sum(item_vars)/total_var), 3) if total_var else 0

    # ── Compute all stats ──────────────────────────────────────────────────────
    def _compute(self):
        r = self.raw

        # Per-question stats
        self.q_stats = {}
        sections = {
            "Q6":"BI","Q7":"BI","Q8":"BI","Q9":"BI","Q10":"BI",
            "Q11":"PA","Q12":"PA","Q13":"PA","Q14":"PA","Q15":"PA",
            "Q16":"Trust","Q17":"Trust","Q18":"Ease","Q19":"Ease",
            "Q20":"Ethics","Q21":"Ethics",
            "Q22":"Success","Q23":"Success","Q24":"Success","Q25":"Success",
        }
        for i, qk in enumerate(Q_LABELS):
            col_idx = 6 + i
            vals = self._val(col_idx)
            self.q_stats[qk] = dict(
                label    = Q_LABELS[qk],
                section  = sections.get(qk,"Other"),
                mean     = self._mean(vals),
                agree    = self._agree(vals),
                neutral  = self._neutral(vals),
                disagree = self._disagree(vals),
                n        = len(vals),
                raw      = vals,
            )

        # Composite means per respondent
        def row_mean(cols):
            out = []
            for row in r:
                try:
                    vals = [float(row[c]) for c in cols]
                    out.append(sum(vals)/len(vals))
                except: pass
            return out

        self.bi_scores  = row_mean(BI_COLS)
        self.pa_scores  = row_mean(PA_COLS)
        self.tr_scores  = row_mean(TRUST_COLS)
        self.suc_scores = row_mean(SUCCESS_COLS)

        self.bi_mean   = self._mean(self.bi_scores)
        self.pa_mean   = self._mean(self.pa_scores)
        self.tr_mean   = self._mean(self.tr_scores)
        self.suc_mean  = self._mean(self.suc_scores)
        self.eth_mean  = self._mean(row_mean(ETHICS_COLS))

        # Cronbach
        def matrix(cols):
            m = []
            for row in r:
                try: m.append([float(row[c]) for c in cols])
                except: pass
            return m

        self.cronbach = dict(
            bi      = self._cronbach(matrix(BI_COLS)),
            pa      = self._cronbach(matrix(PA_COLS)),
            trust   = self._cronbach(matrix(TRUST_COLS)),
            success = self._cronbach(matrix(SUCCESS_COLS)),
        )

        # Correlations
        self.corr = dict(
            bi_success    = self._pearson(self.bi_scores, self.suc_scores),
            pa_success    = self._pearson(self.pa_scores, self.suc_scores),
            trust_success = self._pearson(self.tr_scores, self.suc_scores),
            trust_pa      = self._pearson(self.tr_scores, self.pa_scores),
            bi_pa         = self._pearson(self.bi_scores, self.pa_scores),
        )

        # Demographics
        self.roles   = dict(Counter(self._str_col(1)).most_common())
        self.exp     = dict(Counter(self._str_col(2)).most_common())
        self.method  = dict(Counter(self._str_col(3)).most_common())
        self.org     = dict(Counter(self._str_col(4)).most_common())
        self.sector  = dict(Counter(self._str_col(5)).most_common())
        self.freq    = dict(Counter(self._str_col(26)).most_common())

        # Text responses
        self.barriers = [x for x in self._str_col(27)
                         if len(x)>5 and x.upper() not in ['NA','N/A','NONE','NO PROBLEM']]
        self.comments = [x for x in self._str_col(28)
                         if len(x)>10 and x.upper() not in ['NA','N/A','NONE','NOT APPLICABLE']]

        # Adoption gap
        never_rare = sum(v for k,v in self.freq.items() if k in ['Never','Rarely'])
        freq_vfreq = sum(v for k,v in self.freq.items() if 'Frequent' in k)
        self.rarely_never_pct = round(never_rare/self.n*100, 1)
        self.freq_pct         = round(freq_vfreq/self.n*100, 1)
        self.bi_agree_pct     = self.q_stats["Q6"]["agree"]
        self.pa_agree_pct     = self.q_stats["Q11"]["agree"]
        self.privacy_pct      = self.q_stats["Q21"]["agree"]
        self.bias_pct         = self.q_stats["Q20"]["agree"]

        # Build a respondent-level DataFrame
        self.df = self._build_df()

    def _build_df(self):
        records = []
        for row in self.raw:
            try:
                rec = dict(
                    Role    = str(row[1]).strip() if row[1] else "",
                    Experience = str(row[2]).strip() if row[2] else "",
                    Methodology = str(row[3]).strip() if row[3] else "",
                    Org_Size = str(row[4]).strip() if row[4] else "",
                    Sector   = str(row[5]).strip() if row[5] else "",
                )
                for i,qk in enumerate(Q_LABELS):
                    try: rec[qk] = float(self.raw[self.raw.index(row)][6+i])
                    except: rec[qk] = None
                rec["Q26_Freq"] = str(row[26]).strip() if row[26] else ""
                rec["Q27_Barrier"] = str(row[27]).strip() if row[27] else ""
                rec["Q28_Comment"] = str(row[28]).strip() if row[28] else ""
                records.append(rec)
            except: pass
        return pd.DataFrame(records)
